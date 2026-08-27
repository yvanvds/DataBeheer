#!/usr/bin/env python3
"""Genereert het lesmateriaal voor het Power BI-introductiehoofdstuk (issue #8).

Gebruik:
    python scripts/generate_powerbi_assets.py

Vereist: pip install openpyxl

Het script leest book/_static/db/webshop.db (zelf deterministisch, zie
generate_webshop_db.py) en bevat geen randomness: elke run produceert exact
dezelfde bestanden.

Output:
- book/_static/excel/webshop_verkopen.xlsx
    Excel-export van de webshop: tabblad "Verkopen" (een platte tabel met alle
    orderregels van niet-geannuleerde bestellingen) en tabblad "Voorraad"
    (de productcatalogus met voorraad). Dit is het bestand dat leerlingen
    downloaden voor de klassikale les en de nabouw-challenges.
- book/figures/powerbi_challenge1.svg
- book/figures/powerbi_challenge2.svg
- book/figures/powerbi_challenge3.svg
    Afbeeldingen van de afgewerkte dashboards die leerlingen nabouwen. De
    cijfers in de afbeeldingen komen rechtstreeks uit de database, zodat
    leerlingen hun eigen resultaat exact kunnen controleren.

Onderaan staan asserts die de controlegetallen uit het hoofdstuk
(book/chapters/BIG_DATA/02_Power_BI.ipynb) bewaken: verander de data, en de
asserts vertellen je welke getallen in de tekst moeten worden aangepast.
"""

import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "book" / "_static" / "db" / "webshop.db"
XLSX_PATH = ROOT / "book" / "_static" / "excel" / "webshop_verkopen.xlsx"
FIG_DIR = ROOT / "book" / "figures"

# Standaardkleuren van Power BI (het thema dat leerlingen te zien krijgen).
PALETTE = ["#118DFF", "#12239E", "#E66C37", "#6B007B", "#E044A7",
           "#744EC2", "#D9B300", "#D64550", "#197278"]
BLUE = "#118DFF"
ORANGE = "#E66C37"

FONT = "Segoe UI, sans-serif"
INK = "#252423"        # primaire tekstkleur
INK_MUTED = "#605E5C"  # astitels, aslabels
GRID = "#E8E6E4"
CARD_BORDER = "#E1DFDD"
CANVAS_BG = "#F6F6F6"

# Verwachte controlegetallen; deze staan ook letterlijk in het hoofdstuk.
EXPECTED = {
    "omzet_totaal": 1128345.39,
    "stuks_totaal": 10629,
    "omzet_2025": 612739.43,
    "voorraadwaarde_totaal": 186320.50,
    "voorraadwaarde_computers": 75497.00,
    "omzet_2024": 393843.54,
    "bestellingen_2024": 989,
    "omzet_dec_2024": 98334.52,
    "top_product_2024": ("Cable Kit USB-C", 315),
}


# ---------------------------------------------------------------------------
# Nederlandse notatie
# ---------------------------------------------------------------------------

def nl_int(n) -> str:
    return f"{int(round(n)):,}".replace(",", ".")


def nl_eur(v, dec=2) -> str:
    s = f"{v:,.{dec}f}"
    s = s.replace(",", "@").replace(".", ",").replace("@", ".")
    return f"€ {s}"


def nl_pct(v, dec=1) -> str:
    return f"{v:.{dec}f}%".replace(".", ",")


def esc(s: str) -> str:
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


# ---------------------------------------------------------------------------
# Data ophalen
# ---------------------------------------------------------------------------

def fetch_data():
    con = sqlite3.connect(DB_PATH)
    verkopen = con.execute("""
        SELECT o.order_id,
               o.order_date,
               c.first_name || ' ' || c.last_name AS klant,
               CASE WHEN c.city = 'gent' THEN 'Gent' ELSE c.city END AS stad,
               p.category,
               p.name,
               ol.quantity,
               ol.unit_price,
               ROUND(ol.quantity * ol.unit_price, 2) AS regelbedrag
        FROM order_lines ol
        JOIN orders o    ON o.order_id = ol.order_id
        JOIN customers c ON c.customer_id = o.customer_id
        JOIN products p  ON p.product_id = ol.product_id
        WHERE o.status != 'cancelled'
        ORDER BY o.order_id, p.name
    """).fetchall()
    voorraad = con.execute("""
        SELECT name, category, unit_price, stock
        FROM products
        ORDER BY category, name
    """).fetchall()
    con.close()
    return verkopen, voorraad


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def write_xlsx(verkopen, voorraad):
    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Verkopen"
    headers = ["Bestelnummer", "Datum", "Jaar", "Maand", "Klant", "Stad",
               "Categorie", "Product", "Aantal", "Eenheidsprijs", "Regelbedrag"]
    ws.append(headers)
    for c in ws[1]:
        c.font = bold
    for oid, odate, klant, stad, cat, prod, qty, price, bedrag in verkopen:
        d = date.fromisoformat(odate)
        ws.append([oid, d, d.year, d.month, klant, stad, cat, prod,
                   qty, price, bedrag])
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        row[0].number_format = "DD/MM/YYYY"
    for col in ("J", "K"):
        for row in ws.iter_rows(min_row=2, min_col=ws[f"{col}1"].column,
                                max_col=ws[f"{col}1"].column):
            row[0].number_format = "#,##0.00"
    widths = [13, 12, 6, 7, 20, 14, 12, 20, 8, 13, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Voorraad")
    ws2.append(["Product", "Categorie", "Eenheidsprijs", "Voorraad"])
    for c in ws2[1]:
        c.font = bold
    for name, cat, price, stock in voorraad:
        ws2.append([name, cat, price, stock])
    for row in ws2.iter_rows(min_row=2, min_col=3, max_col=3):
        row[0].number_format = "#,##0.00"
    for i, w in enumerate([20, 12, 13, 9], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # Vaste tijdstempels zodat de output byte-voor-byte reproduceerbaar is.
    fixed = datetime(2026, 1, 1)
    wb.properties.created = fixed
    wb.properties.modified = fixed

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)


# ---------------------------------------------------------------------------
# SVG-tekenprimitieven (nagebootste Power BI-visuals)
# ---------------------------------------------------------------------------

class Svg:
    def __init__(self, w, h):
        self.w, self.h = w, h
        self.parts = [
            f'<svg xmlns="http://www.w3.org/2000/svg" width="{w}" height="{h}" '
            f'viewBox="0 0 {w} {h}" font-family="{FONT}">',
            f'<rect width="{w}" height="{h}" fill="{CANVAS_BG}"/>',
        ]

    def add(self, s):
        self.parts.append(s)

    def rect(self, x, y, w, h, fill, rx=0, stroke=None, sw=1):
        s = f'<rect x="{x:g}" y="{y:g}" width="{w:g}" height="{h:g}" fill="{fill}"'
        if rx:
            s += f' rx="{rx}"'
        if stroke:
            s += f' stroke="{stroke}" stroke-width="{sw}"'
        self.add(s + "/>")

    def line(self, x1, y1, x2, y2, stroke, sw=1):
        self.add(f'<line x1="{x1:g}" y1="{y1:g}" x2="{x2:g}" y2="{y2:g}" '
                 f'stroke="{stroke}" stroke-width="{sw}"/>')

    def text(self, x, y, s, size=11, fill=INK, anchor="start", weight=None,
             style=None):
        attrs = f'x="{x:g}" y="{y:g}" font-size="{size}" fill="{fill}"'
        if anchor != "start":
            attrs += f' text-anchor="{anchor}"'
        if weight:
            attrs += f' font-weight="{weight}"'
        if style:
            attrs += f' {style}'
        self.add(f'<text {attrs}>{esc(s)}</text>')

    def save(self, path: Path):
        self.add("</svg>")
        path.write_text("\n".join(self.parts) + "\n", encoding="utf-8")


def visual_card(svg, x, y, w, h, title=None):
    """Witte visual-kaart met optionele titel; geeft de binnenruimte terug."""
    svg.rect(x, y, w, h, "#FFFFFF", rx=6, stroke=CARD_BORDER)
    top = y
    if title:
        svg.text(x + 14, y + 24, title, size=13, weight="600")
        top = y + 32
    return x + 14, top, w - 28, h - (top - y) - 14


def kpi_card(svg, x, y, w, h, value, label):
    svg.rect(x, y, w, h, "#FFFFFF", rx=6, stroke=CARD_BORDER)
    cx = x + w / 2
    svg.text(cx, y + h / 2 + 4, value, size=30, weight="600", anchor="middle")
    svg.text(cx, y + h / 2 + 30, label, size=12, fill=INK_MUTED, anchor="middle")


def y_axis(svg, px, py, pw, ph, ymax, step, fmt):
    """Horizontale rasterlijnen + labels; geeft schaalfunctie terug."""
    def sy(v):
        return py + ph - (v / ymax) * ph
    v = 0
    while v <= ymax + 1e-9:
        yy = sy(v)
        svg.line(px, yy, px + pw, yy, GRID)
        svg.text(px - 8, yy + 3.5, fmt(v), size=10.5, fill=INK_MUTED,
                 anchor="end")
        v += step
    return sy


def column_chart(svg, x, y, w, h, title, cats, vals, ymax, step,
                 tick_fmt, color=BLUE, data_fmt=None):
    ix, iy, iw, ih = visual_card(svg, x, y, w, h, title)
    lm = 64
    px, py = ix + lm, iy + (18 if data_fmt else 8)
    pw, ph = iw - lm, ih - (18 if data_fmt else 8) - 22
    sy = y_axis(svg, px, py, pw, ph, ymax, step, tick_fmt)
    slot = pw / len(cats)
    bw = slot * 0.55
    for i, (cat, val) in enumerate(zip(cats, vals)):
        bx = px + i * slot + (slot - bw) / 2
        by = sy(val)
        svg.rect(bx, by, bw, py + ph - by, color, rx=2)
        svg.text(px + i * slot + slot / 2, py + ph + 16, cat, size=10.5,
                 fill=INK_MUTED, anchor="middle")
        if data_fmt:
            svg.text(px + i * slot + slot / 2, by - 6, data_fmt(val),
                     size=11, anchor="middle")


def barh_chart(svg, x, y, w, h, title, cats, vals, color=BLUE, data_fmt=None):
    ix, iy, iw, ih = visual_card(svg, x, y, w, h, title)
    lm = max(len(c) for c in cats) * 6.4 + 12
    rm = (max(len(data_fmt(v)) for v in vals) * 6.4 + 12) if data_fmt else 8
    px, py = ix + lm, iy + 6
    pw, ph = iw - lm - rm, ih - 12
    vmax = max(vals)
    slot = ph / len(cats)
    bh = slot * 0.58
    for i, (cat, val) in enumerate(zip(cats, vals)):
        by = py + i * slot + (slot - bh) / 2
        bw = (val / vmax) * pw
        svg.rect(px, by, bw, bh, color, rx=2)
        svg.text(px - 8, by + bh / 2 + 3.5, cat, size=10.5, fill=INK_MUTED,
                 anchor="end")
        if data_fmt:
            svg.text(px + bw + 6, by + bh / 2 + 3.5, data_fmt(val), size=10.5)
    svg.line(px, py, px, py + ph, "#C8C6C4")


def line_chart(svg, x, y, w, h, title, labels, vals, ymax, step, tick_fmt,
               color=BLUE, peak_fmt=None):
    ix, iy, iw, ih = visual_card(svg, x, y, w, h, title)
    lm = 64
    px, py = ix + lm, iy + 16
    pw, ph = iw - lm, ih - 16 - 22
    sy = y_axis(svg, px, py, pw, ph, ymax, step, tick_fmt)
    n = len(vals)
    xs = [px + pw * i / (n - 1) for i in range(n)]
    pts = " ".join(f"{xs[i]:.1f},{sy(v):.1f}" for i, v in enumerate(vals))
    svg.add(f'<polyline points="{pts}" fill="none" stroke="{color}" '
            f'stroke-width="2.5" stroke-linejoin="round"/>')
    for i, v in enumerate(vals):
        svg.add(f'<circle cx="{xs[i]:.1f}" cy="{sy(v):.1f}" r="3.5" '
                f'fill="{color}" stroke="#FFFFFF" stroke-width="1.5"/>')
        svg.text(xs[i], py + ph + 16, labels[i], size=10.5, fill=INK_MUTED,
                 anchor="middle")
    if peak_fmt:
        imax = vals.index(max(vals))
        svg.text(xs[imax] - 4, sy(vals[imax]) - 10, peak_fmt(vals[imax]),
                 size=11, anchor="end", weight="600")


def pie_chart(svg, x, y, w, h, title, cats, vals, min_pct_label=6.0):
    from math import cos, pi, sin
    ix, iy, iw, ih = visual_card(svg, x, y, w, h, title)
    total = sum(vals)
    r = min(ih, iw - 190) / 2 - 4
    cx, cy = ix + r + 12, iy + ih / 2
    a0 = -pi / 2
    for i, val in enumerate(vals):
        frac = val / total
        a1 = a0 + frac * 2 * pi
        large = 1 if frac > 0.5 else 0
        x0, y0 = cx + r * cos(a0), cy + r * sin(a0)
        x1, y1 = cx + r * cos(a1), cy + r * sin(a1)
        svg.add(f'<path d="M{cx:.1f},{cy:.1f} L{x0:.1f},{y0:.1f} '
                f'A{r:.1f},{r:.1f} 0 {large} 1 {x1:.1f},{y1:.1f} Z" '
                f'fill="{PALETTE[i % len(PALETTE)]}" stroke="#FFFFFF" '
                f'stroke-width="2"/>')
        pct = frac * 100
        if pct >= min_pct_label:
            am = (a0 + a1) / 2
            lx, ly = cx + 0.62 * r * cos(am), cy + 0.62 * r * sin(am)
            svg.text(lx, ly + 3.5, nl_pct(pct), size=11, fill="#FFFFFF",
                     anchor="middle", weight="600")
        a0 = a1
    # Legende rechts, in vaste (aflopende) volgorde
    lx = cx + r + 26
    ly = iy + ih / 2 - len(cats) * 9 + 4
    for i, (cat, val) in enumerate(zip(cats, vals)):
        yy = ly + i * 18
        svg.rect(lx, yy - 8, 10, 10, PALETTE[i % len(PALETTE)], rx=2)
        svg.text(lx + 16, yy + 1, f"{cat} — {nl_pct(val / total * 100)}",
                 size=10.5, fill=INK_MUTED)


def slicer(svg, x, y, w, h, title, items, selected=()):
    ix, iy, iw, ih = visual_card(svg, x, y, w, h, None)
    svg.text(ix, iy + 18, title, size=12, weight="600", fill=INK_MUTED)
    for i, item in enumerate(items):
        yy = iy + 34 + i * 26
        box = 13
        if item in selected:
            svg.rect(ix + 2, yy, box, box, BLUE, rx=2)
            svg.add(f'<path d="M{ix + 4.5:.1f},{yy + 6.5:.1f} l3,3.5 l5,-6.5" '
                    f'fill="none" stroke="#FFFFFF" stroke-width="2"/>')
            svg.text(ix + 24, yy + 11, str(item), size=12, weight="600")
        else:
            svg.rect(ix + 2, yy, box, box, "#FFFFFF", rx=2,
                     stroke=INK_MUTED, sw=1.2)
            svg.text(ix + 24, yy + 11, str(item), size=12)


def table_visual(svg, x, y, w, h, title, headers, rows, aligns, total_row=None,
                 scrollbar=False):
    ix, iy, iw, ih = visual_card(svg, x, y, w, h, title)
    ncol = len(headers)
    col_w = [0.5, 0.25, 0.25] if ncol == 3 else [1 / ncol] * ncol
    xs = [ix]
    for cw in col_w:
        xs.append(xs[-1] + cw * (iw - (14 if scrollbar else 0)))
    row_h = 24
    yy = iy + 6

    def cell(cx0, cx1, text, align, ypos, weight=None, fill=INK):
        if align == "right":
            svg.text(cx1 - 8, ypos, text, size=11, anchor="end",
                     weight=weight, fill=fill)
        else:
            svg.text(cx0 + 4, ypos, text, size=11, weight=weight, fill=fill)

    for j, htxt in enumerate(headers):
        cell(xs[j], xs[j + 1], htxt, aligns[j], yy + 15, weight="600")
    svg.line(ix, yy + row_h - 2, xs[-1], yy + row_h - 2, INK)
    yy += row_h
    for row in rows:
        for j, val in enumerate(row):
            cell(xs[j], xs[j + 1], val, aligns[j], yy + 16)
        svg.line(ix, yy + row_h, xs[-1], yy + row_h, GRID)
        yy += row_h
    if total_row:
        for j, val in enumerate(total_row):
            cell(xs[j], xs[j + 1], val, aligns[j], yy + 16, weight="600")
        svg.line(ix, yy, xs[-1], yy, INK)
    if scrollbar:
        sx = ix + iw - 8
        svg.rect(sx, iy + 6, 6, ih - 12, "#F3F2F1", rx=3)
        svg.rect(sx, iy + 6, 6, (ih - 12) * 0.35, "#C8C6C4", rx=3)


# ---------------------------------------------------------------------------
# De drie dashboards
# ---------------------------------------------------------------------------

def eur_tick(v):
    return "€ 0" if v == 0 else f"€ {nl_int(v)}"


def challenge1(agg):
    svg = Svg(1180, 640)
    kpi_card(svg, 20, 20, 270, 130, nl_int(agg["stuks_totaal"]),
             "Totaal verkochte stuks")
    cats, vals = zip(*agg["omzet_per_cat"])
    column_chart(svg, 310, 20, 850, 290, "Omzet per categorie",
                 cats, vals, ymax=350000, step=100000, tick_fmt=eur_tick)
    pcats, pvals = zip(*agg["stuks_per_cat"])
    pie_chart(svg, 20, 330, 560, 290, "Verkochte stuks per categorie",
              pcats, pvals)
    jaren, jvals = zip(*agg["omzet_per_jaar"])
    column_chart(svg, 600, 330, 560, 290, "Omzet per jaar",
                 jaren, jvals, ymax=700000, step=200000, tick_fmt=eur_tick,
                 data_fmt=lambda v: nl_eur(v))
    svg.save(FIG_DIR / "powerbi_challenge1.svg")


def challenge2(agg):
    svg = Svg(1180, 700)
    svg.text(24, 46, "Voorraadrapport webshop", size=24, weight="600")
    kpi_card(svg, 20, 70, 270, 130, nl_eur(agg["voorraadwaarde_totaal"]),
             "Totale voorraadwaarde")
    cats, vals = zip(*agg["voorraadwaarde_per_cat"])
    barh_chart(svg, 310, 70, 850, 330, "Voorraadwaarde per categorie",
               cats, vals, color=ORANGE, data_fmt=lambda v: nl_eur(v))
    rows = [(name, nl_int(stock), nl_eur(waarde))
            for name, stock, waarde in agg["top10_voorraadwaarde"]]
    table_visual(svg, 20, 420, 1140, 260,
                 "Producten volgens voorraadwaarde",
                 ["Product", "Voorraad", "Voorraadwaarde"],
                 rows[:8], ["left", "right", "right"],
                 total_row=("Totaal", nl_int(agg["voorraad_stuks_totaal"]),
                            nl_eur(agg["voorraadwaarde_totaal"])),
                 scrollbar=True)
    svg.save(FIG_DIR / "powerbi_challenge2.svg")


def challenge3(agg):
    svg = Svg(1180, 700)
    slicer(svg, 20, 20, 220, 130, "Jaar", [2023, 2024, 2025], selected=(2024,))
    slicer(svg, 20, 170, 220, 510, "Categorie",
           [c for c, _ in agg["omzet_per_cat"]])
    kpi_card(svg, 260, 20, 440, 130, nl_eur(agg["omzet_2024"]), "Omzet")
    kpi_card(svg, 720, 20, 440, 130, nl_int(agg["bestellingen_2024"]),
             "Aantal bestellingen")
    labels = [str(m) for m in range(1, 13)]
    line_chart(svg, 260, 170, 900, 250, "Omzet per maand",
               labels, agg["omzet_per_maand_2024"], ymax=100000, step=25000,
               tick_fmt=eur_tick, peak_fmt=lambda v: nl_eur(v))
    cats, vals = zip(*agg["top10_stuks_2024"])
    barh_chart(svg, 260, 440, 900, 240,
               "Top 10 producten volgens verkochte stuks",
               cats, vals, data_fmt=lambda v: nl_int(v))
    svg.save(FIG_DIR / "powerbi_challenge3.svg")


# ---------------------------------------------------------------------------
# Aggregaties + controle
# ---------------------------------------------------------------------------

def aggregate(verkopen, voorraad):
    from collections import defaultdict
    omzet_cat = defaultdict(float)
    stuks_cat = defaultdict(int)
    omzet_jaar = defaultdict(float)
    omzet_2024_maand = defaultdict(float)
    orders_2024 = set()
    stuks_2024_prod = defaultdict(int)
    totaal_omzet = 0.0
    totaal_stuks = 0
    for oid, odate, _klant, _stad, cat, prod, qty, _price, bedrag in verkopen:
        jaar, maand = int(odate[:4]), int(odate[5:7])
        omzet_cat[cat] += bedrag
        stuks_cat[cat] += qty
        omzet_jaar[jaar] += bedrag
        totaal_omzet += bedrag
        totaal_stuks += qty
        if jaar == 2024:
            omzet_2024_maand[maand] += bedrag
            orders_2024.add(oid)
            stuks_2024_prod[prod] += qty

    vw_cat = defaultdict(float)
    vw_rows = []
    vw_totaal = 0.0
    vs_totaal = 0
    for name, cat, price, stock in voorraad:
        w = round(price * stock, 2)
        vw_cat[cat] += w
        vw_rows.append((name, stock, w))
        vw_totaal += w
        vs_totaal += stock
    vw_rows.sort(key=lambda r: -r[2])

    by_desc = lambda d: sorted(d.items(), key=lambda kv: -kv[1])
    return {
        "omzet_totaal": round(totaal_omzet, 2),
        "stuks_totaal": totaal_stuks,
        "omzet_per_cat": [(c, round(v, 2)) for c, v in by_desc(omzet_cat)],
        "stuks_per_cat": by_desc(stuks_cat),
        "omzet_per_jaar": [(j, round(omzet_jaar[j], 2))
                           for j in sorted(omzet_jaar)],
        "voorraadwaarde_totaal": round(vw_totaal, 2),
        "voorraadwaarde_per_cat": [(c, round(v, 2)) for c, v in by_desc(vw_cat)],
        "top10_voorraadwaarde": vw_rows[:10],
        "voorraad_stuks_totaal": vs_totaal,
        "omzet_2024": round(sum(omzet_2024_maand.values()), 2),
        "bestellingen_2024": len(orders_2024),
        "omzet_per_maand_2024": [round(omzet_2024_maand[m], 2)
                                 for m in range(1, 13)],
        "top10_stuks_2024": sorted(stuks_2024_prod.items(),
                                   key=lambda kv: (-kv[1], kv[0]))[:10],
    }


def verify(agg, verkopen):
    assert agg["omzet_totaal"] == EXPECTED["omzet_totaal"], agg["omzet_totaal"]
    assert agg["stuks_totaal"] == EXPECTED["stuks_totaal"], agg["stuks_totaal"]
    assert dict(agg["omzet_per_jaar"])[2025] == EXPECTED["omzet_2025"]
    assert agg["voorraadwaarde_totaal"] == EXPECTED["voorraadwaarde_totaal"]
    assert dict(agg["voorraadwaarde_per_cat"])["Computers"] == \
        EXPECTED["voorraadwaarde_computers"]
    assert agg["omzet_2024"] == EXPECTED["omzet_2024"], agg["omzet_2024"]
    assert agg["bestellingen_2024"] == EXPECTED["bestellingen_2024"]
    assert agg["omzet_per_maand_2024"][11] == EXPECTED["omzet_dec_2024"]
    assert agg["top10_stuks_2024"][0] == EXPECTED["top_product_2024"]
    assert all(stad != "gent" for _o, _d, _k, stad, *_ in verkopen)

    print(f"OK: {len(verkopen)} verkoopregels geexporteerd")
    print(f"    totale omzet:          {nl_eur(agg['omzet_totaal'])}")
    print(f"    verkochte stuks:       {nl_int(agg['stuks_totaal'])}")
    print(f"    omzet 2025:            {nl_eur(dict(agg['omzet_per_jaar'])[2025])}")
    print(f"    voorraadwaarde:        {nl_eur(agg['voorraadwaarde_totaal'])}")
    print(f"    omzet 2024:            {nl_eur(agg['omzet_2024'])}")
    print(f"    bestellingen 2024:     {agg['bestellingen_2024']}")
    print(f"    piekmaand 2024 (dec):  {nl_eur(agg['omzet_per_maand_2024'][11])}")
    print(f"    bestand: {XLSX_PATH} ({XLSX_PATH.stat().st_size / 1024:.0f} KiB)")


def main():
    verkopen, voorraad = fetch_data()
    agg = aggregate(verkopen, voorraad)
    write_xlsx(verkopen, voorraad)
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    challenge1(agg)
    challenge2(agg)
    challenge3(agg)
    verify(agg, verkopen)


if __name__ == "__main__":
    main()
